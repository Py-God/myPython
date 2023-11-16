import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

get_column_letter(1)
'A'
get_column_letter(2)
'B'
get_column_letter(27)
'AA'
get_column_letter(900)
'AHP'

wb = openpyxl.load_workbook('C:\\Users\\user\\Desktop\\PYTHON\\example.xlsx')
sheet = wb['Sheet1']
get_column_letter(sheet.max_column)
'C'
column_index_from_string('A')
1
column_index_from_string('AA')
27
