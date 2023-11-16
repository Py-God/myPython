import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active
sheet['A2'] = 100

# change font styles of a particular cell
cell = sheet['A2']
cell.font = Font(name='Calibri',
                 size=20,
                 bold=True,
                 italic=True,
                 underline='single',
                 color='008000')

# fill cell color
cell.fill = PatternFill(fill_type='solid',
                        start_color='FFFF00',
                        end_color='FFFF00')

# resize cell
sheet.row_dimensions[2].height = 40
sheet.column_dimensions['A'].width = 20

# change styles of many cells at once
for row in range(3,8):
    for col in range(3,6):
        sheet.cell(row,col).fill = PatternFill(fill_type='solid',
                                               start_color='00FF6600',
                                               end_color='00ff6600')
        sheet.cell(row,col).border = Border(left=Side(border_style='double', color='FF000000'),
                                            right=Side(border_style='double', color='FF000000'),
                                            top=Side(border_style='double', color='FF000000'),
                                            bottom=Side(border_style='double', color='FF000000'))

        sheet.row_dimensions[row].height = 80
        col_letter = get_column_letter(col)
        sheet.column_dimensions(col_letter).width = 20
wb.save('styles.xlsx')