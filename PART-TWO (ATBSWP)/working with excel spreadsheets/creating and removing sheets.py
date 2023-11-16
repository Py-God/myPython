import openpyxl

wb = openpyxl.Workbook()
wb.sheetnames
['Sheet']
wb.create_sheet()
'<Worksheet "Sheet1">'
wb.sheetnames
['Sheet', 'Sheet1']
wb.create_sheet(index=0, title='First Sheet')
'<Worksheet "First Sheet">'
wb.sheetnames
['First Sheet', 'Sheet', 'Sheet1']
wb.create_sheet(index=2, title='Middle Sheet')
'<Worksheet "Middle Sheet">'
wb.sheetnames
['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']

wb.sheetnames
['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']
del wb['Middle Sheet']
wb.sheetnames
['First Sheet', 'Sheet', 'Sheet1']
del wb['Sheet1']
wb.sheetnames
['First Sheet', 'Sheet']
