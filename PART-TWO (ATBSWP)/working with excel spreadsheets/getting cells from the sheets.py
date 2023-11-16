import openpyxl

wb = openpyxl.load_workbook('C:\\Users\\user\\Desktop\\PYTHON\\example.xlsx')

sheet = wb['Sheet1']
'<Worksheet "Sheet1">'

sheet['A1'].value
'datetime.datetime(2015, 4, 5, 13, 34, 2)'
c = sheet['B1']
c.value
'Apples'

'Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value
'Row 1, Column 2 is Apples'
'Cell ' + c.coordinate + ' is ' + c.value

'Cell B1 is Apples'
sheet['C1'].value
73
sheet.cell(row=1, column=2)
"<Cell 'Sheet1'.B1>"
sheet.cell(row=1, column=2).value
'Apples'
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
'''
1 Apples
3 Pears
5 Apples
7 Strawberries
'''

sheet = wb['Sheet1']
sheet.max_row
7
sheet.max_column
3
