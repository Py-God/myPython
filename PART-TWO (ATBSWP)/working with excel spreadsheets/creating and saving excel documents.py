Python 3.10.0 (tags/v3.10.0:b494f59, Oct  4 2021, 19:00:18) [MSC v.1929 64 bit (AMD64)] on win32
Type "help", "copyright", "credits" or "license()" for more information.
import openpyxl
wb = openpyxl.Workbook()
wb.sheetnames
['Sheet']
sheet = wb.active
sheet.title
'Sheet'
sheet.title = 'Spam Bacon Eggs Sheet'
wb.sheetnames
['Spam Bacon Eggs Sheet']

wb = openpyxl.load_workbook('C:\\Users\\user\\Desktop\\PYTHON\\example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx')
